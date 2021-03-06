import base64
import csv
import json
import os
import shutil
import tempfile
import zipfile
from datetime import datetime
from errno import ENOTDIR
from itertools import islice
from shlex import split
from subprocess import PIPE, Popen
from uuid import uuid4

import pandas as pd
from django.conf import settings
from django.core.files.storage import FileSystemStorage
from openpyxl import load_workbook
from rest_framework import exceptions

from pyplan.pyplan.common.baseService import BaseService
from pyplan.pyplan.companies.models import Company
from pyplan.pyplan.preference.models import Preference
from pyplan.pyplan.user_company_preference.models import UserCompanyPreference

from .classes.fileEntry import FileEntry, eFileTypes
from .classes.fileEntryData import (FileEntryData, eSpecialFileType,
                                    eSpecialFolder)


class FileManagerService(BaseService):

    optimizable_templates = [".xls", ".xlsx", ".xlsm", ".xlsb"]

    def getMainFolders(self):
        """
        list:
        Return a list of all folders.
        """

        result = list()
        company_code = self.client_session.company_code
        # User Workspace
        result.append(
            FileEntry(
                text="My Workspace",
                type=eFileTypes.NOTHING,
                data=FileEntryData(
                    fullPath="",
                    specialFolderType=eSpecialFolder.MY_FOLDER
                )
            )
        )

        # if self.current_user.has_perm("pyplan.view_model_path_root"):
        #     result.append(
        #         FileEntry(
        #             text="Root",
        #             type=eFileTypes.NOTHING,
        #             data=FileEntryData(
        #                 fullPath="",
        #                 specialFolderType=eSpecialFolder.MODELS_PATH
        #             )
        #         )
        #     )

        # if self.current_user.has_perm("pyplan.view_company_root"):
        #     result.append(
        #         FileEntry(
        #             text=self.client_session.companyName,
        #             type=eFileTypes.NOTHING,
        #             data=FileEntryData(
        #                 fullPath=company_code,
        #                 specialFolderType=eSpecialFolder.MODELS_PATH
        #             )
        #         )
        #     )
        # else:
        #     # Special Folders
        #     result.append(
        #         FileEntry(
        #             text="Public",
        #             type=eFileTypes.NOTHING,
        #             data=FileEntryData(
        #                 fullPath=f"{company_code}/Public",
        #                 specialFolderType=eSpecialFolder.PUBLIC
        #             )
        #         )
        #     )

        return result

    def getFoldersAndFiles(self, folder=''):
        """
        list:
        Return a list of all folders.
        """
        storage = FileSystemStorage(
            os.path.join(settings.MEDIA_ROOT, 'models'))
        result = list()

        base_path = ""
        if folder.startswith("/"):
            base_path = folder[1:]
        elif folder:
            base_path = f"{folder}/"

        items = storage.listdir(base_path)

        denied_folders = []
        if not self.current_user.has_perm('pyplan.change_group_permissions'):
            denied_folders = self._getDeniedFolders()

        # folders
        for item in sorted(items[0], key=str.lower):
            full_path = f"{base_path}{item}"
            if not denied_folders or not item in denied_folders:
                result.append(
                    FileEntry(
                        show=not item.startswith('.'),
                        text=item,
                        type=eFileTypes.MY_FOLDER,
                        data=FileEntryData(
                            fileSize=None,
                            fullPath=full_path,
                            # specialFolderType=eSpecialFolder.MODELS_PATH
                            lastUpdateTime=storage.get_modified_time(
                                full_path),
                        )
                    )
                )
        # files
        for item in sorted(items[1], key=str.lower):
            full_path = f"{base_path}{item}"
            specialFileType = eSpecialFileType.FILE
            lowerItem = item.lower()
            if lowerItem.endswith('.ppl') | lowerItem.endswith('.cbpy') | \
                    lowerItem.endswith('.model') | lowerItem.endswith('.ana'):
                specialFileType = eSpecialFileType.MODEL
            elif lowerItem.endswith('.zip'):
                specialFileType = eSpecialFileType.ZIP

            result.append(
                FileEntry(
                    text=item,
                    type=eFileTypes.PUBLIC,
                    data=FileEntryData(
                        fileSize=storage.size(full_path),
                        fullPath=full_path,
                        extension=full_path[full_path.rfind('.')+1:],
                        specialFileType=specialFileType,
                        lastUpdateTime=storage.get_modified_time(full_path),
                    )
                )
            )
        return result

    def createFolder(self, folder_path, folder_name):
        """
        create:
        Creates a folder inside provided path.
        """
        storage = FileSystemStorage(
            os.path.join(settings.MEDIA_ROOT, 'models'))

        full_path = os.path.join(
            storage.base_location, folder_path, folder_name)

        if storage.exists(full_path):
            raise exceptions.NotAcceptable('Folder already exists')
        else:
            os.mkdir(full_path)
            return os.path.join(folder_path, folder_name)

    def createFile(self, my_file, folder_path, name, chunk):
        file_path = os.path.join(
            settings.MEDIA_ROOT, 'models', folder_path, name)
        # Moves file if it already exists
        if chunk is 0 and os.path.isfile(file_path):
            new_name = f"{name}-{datetime.today().strftime('%Y%m%d-%H:%M:%S')}.old"
            self._copy(
                file_path,
                os.path.join(settings.MEDIA_ROOT, 'models',
                             folder_path, new_name)
            )
            os.remove(file_path)
        # Appends all chunks of this request (chunks of chunks)
        # UI sends multiple requests with multiple chunks each per file
        with open(file_path, 'ab+') as temp_file:
            for chunk in my_file.chunks():
                temp_file.write(chunk)

    def copyFileOrFolder(self, source, destination):
        """
        create:
        Duplicate file or Folder.
        """

        storage = FileSystemStorage(
            os.path.join(settings.MEDIA_ROOT, 'models'))

        return self._linuxCopy(
            f"{storage.base_location}/{source}",
            f"{storage.base_location}/{destination}"
        )

    def ensureUserWorkspace(self):
        storage = FileSystemStorage(
            os.path.join(settings.MEDIA_ROOT, 'models'))

        # User Workspace
        if not storage.exists(storage.base_location):
            os.makedirs(storage.base_location)

    def renameFile(self, source, new_name):
        storage = FileSystemStorage(
            os.path.join(settings.MEDIA_ROOT, 'models'))
        src = f"{storage.base_location}/{source}"
        dest = f"{src[0:src.rfind('/')+1]}{new_name}"
        os.rename(src, dest)
        return dest

    def duplicateFiles(self, sources):
        result = []
        storage = FileSystemStorage(
            os.path.join(settings.MEDIA_ROOT, 'models'))
        for source in sources:
            src = f"{storage.base_location}/{source}"
            dest_path, dest_name = source.rsplit('/', 1)
            dest = f"{storage.base_location}/{dest_path}/Copy 1 of {dest_name}"
            n = 1
            while storage.exists(dest):
                n += 1
                dest = f"{storage.base_location}/{dest_path}/Copy {n} of {dest_name}"

            result.append(self._linuxCopy(src, dest))
        return result

    def moveFiles(self, sources, target):
        result = []
        storage = FileSystemStorage(
            os.path.join(settings.MEDIA_ROOT, 'models'))
        for source in sources:
            src = f"{storage.base_location}/{source}"
            dest_path, dest_name = source.rsplit('/', 1)
            dest = f"{storage.base_location}/{target}/{dest_name}"
            result.append(self.recursive_overwrite(src, dest))
            if os.path.isdir(src):
                shutil.rmtree(src)
            else:
                storage.delete(src)
        return result

    def copyFiles(self, sources, target):
        storage = FileSystemStorage(
            os.path.join(settings.MEDIA_ROOT, 'models'))
        for source in sources:
            src = f"{storage.base_location}/{source}"
            dest_path, dest_name = source.rsplit('/', 1)
            dest = f"{storage.base_location}/{target}/"
            self._linuxCopy(src, dest)
        return True

    def copyToMyWorkspace(self, source):
        storage = FileSystemStorage(
            os.path.join(settings.MEDIA_ROOT, 'models'))

        target = f"{storage.base_location}/{self.client_session.company_code}/{self.current_user.username}"

        src = f"{storage.base_location}/{source}"
        return self._linuxCopy(src, target)

    def deleteFiles(self, sources):
        storage = FileSystemStorage(
            os.path.join(settings.MEDIA_ROOT, 'models'))
        files = []
        for source in sources:
            full_path = f'{storage.base_location}/{source}'
            if not storage.exists(full_path):
                raise exceptions.NotAcceptable(f'File {source} does not exist')
            else:
                files.append(full_path)
        for file_to_delete in files:
            if os.path.isfile(file_to_delete):
                storage.delete(file_to_delete)
            else:
                shutil.rmtree(file_to_delete)

    def download(self, sources):
        storage = FileSystemStorage(
            os.path.join(settings.MEDIA_ROOT, 'models'))

        src_0 = f"{storage.base_location}/{sources[0]}"
        if len(sources) is 1 and os.path.isfile(src_0):
            return open(src_0, 'rb'), os.path.relpath(src_0, os.path.join(src_0, '..'))
        else:
            temp = tempfile.SpooledTemporaryFile()
            with zipfile.ZipFile(temp, 'w', zipfile.ZIP_DEFLATED) as zfobj:
                for source in sources:
                    src = f"{storage.base_location}/{source}"
                    if os.path.isfile(src):
                        zfobj.write(src, os.path.relpath(
                            src, os.path.join(src, '..')))
                    else:
                        self._zipdir(src, zfobj)
                for zfile in zfobj.filelist:
                    zfile.create_system = 0
            temp.seek(0)
            return temp, f"{os.path.relpath(sources[0], os.path.join(sources[0], '..'))}.zip"

    def makeJsonStream(self, json_string: str):
        """
        Returns streamed temp file from json string
        """
        temp = tempfile.SpooledTemporaryFile(mode='w+b')
        temp.write(str.encode(json_string))
        temp.seek(0)
        return temp

    def unzipFile(self, source, target_folder):
        storage = FileSystemStorage(
            os.path.join(settings.MEDIA_ROOT, 'models'))
        src = f"{storage.base_location}/{source}"
        dest = f"{storage.base_location}/{target_folder}"

        # Unzip the file, creating subdirectories as needed
        zfobj = zipfile.ZipFile(src)
        zfobj.extractall(dest)

    def zipFiles(self, sources):
        storage = FileSystemStorage(
            os.path.join(settings.MEDIA_ROOT, 'models'))
        zip_file = f"{storage.base_location}/{sources[0]}.zip"

        if storage.exists(zip_file):
            file_name, file_extension = os.path.splitext(zip_file)
            zip_file = f'{file_name}_{uuid4().hex}{file_extension}'

        with zipfile.ZipFile(zip_file, 'w', zipfile.ZIP_DEFLATED) as zfobj:
            for source in sources:
                src = f"{storage.base_location}/{source}"
                if os.path.isfile(src):
                    zfobj.write(src, os.path.relpath(
                        src, os.path.join(src, '..')))
                else:
                    self._zipdir(src, zfobj)
            for zfile in zfobj.filelist:
                zfile.create_system = 0

    def getHome(self):
        company_id = self.getSession().companyId
        model_path = eSpecialFolder.MODELS_PATH
        res = {}

        filepath = os.path.join(settings.MEDIA_ROOT,
                                'models', 'home.json')
        if os.path.isfile(filepath):
            with open(filepath, "r") as json_file:
                try:
                    res = json.load(json_file)
                    if "tabs" in res:
                        for tab in res["tabs"]:
                            if "folders" in tab:
                                for folder in tab["folders"]:
                                    if "items" in folder:
                                        for item in folder["items"]:
                                            if "image" in item:
                                                image_path = os.path.join(
                                                    settings.MEDIA_ROOT, 'models', item["image"])
                                                if os.path.isfile(image_path):
                                                    with open(image_path, "rb") as f_image:
                                                        item["imagesrc"] = str(
                                                            base64.b64encode(f_image.read()), "utf-8")
                except Exception as ex:
                    raise exceptions.NotAcceptable(ex)
        return res

    def optimizeTemplates(self, sources):
        """Generate csv file for each named range in template for future read.
        """

        if sources:
            preference = Preference.objects.filter(
                code="optimize_templates").first()
            if preference:
                user_company_id = self.client_session.userCompanyId
                c_pref = UserCompanyPreference.objects.filter(
                    user_company_id=user_company_id, preference__code="optimize_templates").first()
                if c_pref:
                    preference.definition = c_pref.definition

            if preference.definition["value"]:
                for template in sources:
                    filename, file_extension = os.path.splitext(template)
                    if file_extension in self.optimizable_templates:
                        template_filename = os.path.join(
                            settings.MEDIA_ROOT, 'models', template)
                        self._generate_csv_from_excel(template_filename)

    # Private

    def _zipdir(self, path, ziph):
        denied_folders = self._getDeniedFolders()
        # ziph is zipfile handle
        for root, dirs, files in os.walk(path):
            # check if folder is not in any department denied folders
            if not denied_folders or not any(list(map(lambda item: item in denied_folders, root.rsplit('/')))):
                for file in files:
                    ziph.write(os.path.join(root, file), os.path.relpath(
                        os.path.join(root, file), os.path.join(path, '..')))

    def _generate_csv_from_excel(self, filename):
        """Generate compressed csv from excel file
        """

        target_dir = os.path.dirname(filename)
        file_name, file_extension = os.path.splitext(filename)
        target_dir = os.path.join(target_dir, file_name)

        if not os.path.isdir(target_dir):
            os.mkdir(target_dir)

        wb = load_workbook(filename, data_only=True, read_only=True)
        for item in wb.defined_names.definedName:
            if not item.is_external and item.type == "RANGE" and item.attr_text and "!$" in item.attr_text:
                target_filename = os.path.join(target_dir, item.name+".pkl")
                if os.path.isfile(target_filename):
                    os.remove(target_filename)

                dests = item.destinations
                for title, coord in dests:
                    if title in wb:
                        ws = wb[title]
                        rangeToRead = ws[coord]
                        if not isinstance(rangeToRead, tuple):
                            rangeToRead = ((rangeToRead,),)

                        nn = 0
                        cols = []
                        values = []
                        for row in rangeToRead:
                            if nn == 0:
                                cols = [str(c.value) for c in row]
                            else:
                                values.append([c.value for c in row])
                            nn += 1
                        nn = 0
                        _finalCols = []
                        for _col in cols:
                            if _col is None:
                                _finalCols.append("Unnamed" + str(nn))
                                nn += 1
                            else:
                                _finalCols.append(_col)
                        df = pd.DataFrame(
                            values, columns=_finalCols).dropna(how="all")
                        df.to_pickle(target_filename, compression='gzip')

    def recursive_overwrite(self, src, dest, ignore=None):
        if os.path.isdir(src):
            if not os.path.isdir(dest):
                os.makedirs(dest)
            files = os.listdir(src)
            if ignore is not None:
                ignored = ignore(src, files)
            else:
                ignored = set()
            for f in files:
                if f not in ignored:
                    self.recursive_overwrite(
                        os.path.join(src, f),
                        os.path.join(dest, f),
                        ignore
                    )
        else:
            shutil.copy(src, dest)
        return dest

    def _copy(self, src, dest):
        try:
            return shutil.copytree(src, dest)
        except OSError as e:
            # If the error was caused because the source wasn't a directory
            if e.errno == ENOTDIR:
                return shutil.copy(src, dest)
            else:
                raise exceptions.NotAcceptable(
                    'Directory not copied. Error: %s' % e)
        except Exception as e:
            raise e

    def _linuxCopy(self, src, dest):
        src_path = src.replace(' ', '\ ')
        dest_path = dest.replace(' ', '\ ')

        # -R, -r, --recursive
        #   copy directories recursively
        # -u, --update
        #   copy only when the SOURCE file is newer
        #   than the destination file or when the
        #   destination file is missing
        # -v, --verbose
        #   explain what is being done

        cmd = f'cp -ruv {src_path} {dest_path}'
        popen = Popen(split(cmd), stdout=PIPE, universal_newlines=True)

        stdout, stderr = popen.communicate()
        if stderr:
            raise exceptions.NotAcceptable(stderr)

        return True
